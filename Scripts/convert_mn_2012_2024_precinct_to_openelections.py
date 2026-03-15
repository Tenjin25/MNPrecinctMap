#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass
class SourceJob:
    year: int
    source_type: str  # xlsx | csv
    source_path: Path
    county_results_path: Path
    output_path: Path


@dataclass
class ContestSpec:
    column: str
    office: str
    district_kind: str
    party: str
    candidate: str


DEFAULT_JOBS = [
    SourceJob(
        year=2012,
        source_type="xlsx",
        source_path=Path("Data/2012_general_precinct_official.xlsx"),
        county_results_path=Path("Data/20121106__mn__general__county.csv"),
        output_path=Path("Data/20121106__mn__general__precinct.csv"),
    ),
    SourceJob(
        year=2014,
        source_type="xlsx",
        source_path=Path("Data/2014_general_precinct_official.xlsx"),
        county_results_path=Path("Data/20141104__mn__general__county.csv"),
        output_path=Path("Data/20141104__mn__general__precinct.csv"),
    ),
    SourceJob(
        year=2016,
        source_type="xlsx",
        source_path=Path("Data/2016_general_precinct_official.xlsx"),
        county_results_path=Path("Data/20161108__mn__general__county.csv"),
        output_path=Path("Data/20161108__mn__general__precinct.csv"),
    ),
    SourceJob(
        year=2018,
        source_type="xlsx",
        source_path=Path("Data/2018_general_precinct_official.xlsx"),
        county_results_path=Path("Data/20181106__mn__general__county.csv"),
        output_path=Path("Data/20181106__mn__general__precinct.csv"),
    ),
    SourceJob(
        year=2024,
        source_type="csv",
        source_path=Path("Data/2024-general-federal-state-results-by-precinct-official - Precinct-Results.csv"),
        county_results_path=Path("Data/20241105__mn__general__county.csv"),
        output_path=Path("Data/20241105__mn__general__precinct.csv"),
    ),
]

PREFIX_CONTEST_MAP = {
    "USSSE": ("U.S. Senate, Unexpired Term", "unexpired"),
    "USPRS": ("President", "na"),
    "USSEN": ("U.S. Senate", "na"),
    "USREP": ("U.S. House", "cong"),
    "MNSEN": ("State Senate", "mnsen"),
    "MNLEG": ("State House", "mnleg"),
    "MNGOV": ("Governor", "na"),
    "MNSOS": ("Secretary of State", "na"),
    "MNAUD": ("State Auditor", "na"),
    "MNAG": ("Attorney General", "na"),
}


def clean(value: str) -> str:
    return (value or "").strip()


def normalize_house_district(value: str) -> str:
    token = clean(value).upper()
    if token in {"", "NA"}:
        return "NA"
    m = re.match(r"^0*([0-9]+)([A-Z]?)$", token)
    if not m:
        return token
    return f"{int(m.group(1))}{m.group(2)}"


def normalize_numeric_district(value: str) -> str:
    token = clean(value)
    if token in {"", "NA"}:
        return "NA"
    try:
        return str(int(float(token)))
    except ValueError:
        return token


def normalize_office(value: str) -> str:
    office = clean(value)
    if office == "Governor & Lt Governor":
        return "Governor"
    if office.upper() == "CONSTITUTIONAL AMENDMENT 1":
        return "Constitutional Amendment 1"
    if office.upper() == "CONSTITUTIONAL AMENDMENT 2":
        return "Constitutional Amendment 2"
    return office


def normalize_district_for_office(office: str, district: str) -> str:
    office = normalize_office(office)
    if office == "U.S. House":
        return normalize_numeric_district(district)
    if office == "State Senate":
        return normalize_numeric_district(district)
    if office == "State House":
        return normalize_house_district(district)
    if office == "U.S. Senate, Unexpired Term":
        return "Unexpired Term"
    return "NA"


def parse_int(value: str) -> int | None:
    token = clean(value).replace(",", "")
    if token == "":
        return None
    try:
        return int(float(token))
    except ValueError:
        return None


def load_rows_from_xlsx(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_name = "Results" if "Results" in wb.sheetnames else "Precinct-Results"
    ws = wb[sheet_name]

    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter)
    headers = [clean(str(h)) if h is not None else "" for h in raw_headers]

    rows: list[dict[str, str]] = []
    for raw in rows_iter:
        row: dict[str, str] = {}
        has_data = False
        for i, header in enumerate(headers):
            if header == "":
                continue
            value = raw[i] if i < len(raw) else None
            if value is None:
                row[header] = ""
                continue
            if isinstance(value, float) and value.is_integer():
                text = str(int(value))
            else:
                text = str(value)
            row[header] = clean(text)
            if row[header] != "":
                has_data = True
        if has_data:
            rows.append(row)
    return headers, rows


def load_rows_from_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    rows: list[dict[str, str]] = []
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        headers = [clean(h) for h in (reader.fieldnames or [])]
        for raw in reader:
            row = {clean(k): clean(v) for k, v in raw.items()}
            rows.append(row)
    return headers, rows


def load_county_candidate_map(path: Path) -> dict[tuple[str, str, str], str]:
    mapping: dict[tuple[str, str, str], str] = {}
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            office = normalize_office(row.get("office", ""))
            district = normalize_district_for_office(office, row.get("district", ""))
            party = clean(row.get("party", ""))
            candidate = clean(row.get("candidate", ""))
            if not office or not party or not candidate:
                continue
            key = (office, district, party)
            if key not in mapping:
                mapping[key] = candidate
    return mapping


def build_contests(headers: list[str]) -> list[ContestSpec]:
    specs: list[ContestSpec] = []
    for header in headers:
        h = clean(header).upper()
        if h == "":
            continue

        if h.startswith("MNCA1"):
            suffix = h.replace("MNCA1", "", 1)
            if suffix in {"YES", "NO"}:
                specs.append(
                    ContestSpec(
                        column=header,
                        office="Constitutional Amendment 1",
                        district_kind="na",
                        party="NP",
                        candidate=suffix,
                    )
                )
            continue

        if h.startswith("MNCA2"):
            suffix = h.replace("MNCA2", "", 1)
            if suffix in {"YES", "NO"}:
                specs.append(
                    ContestSpec(
                        column=header,
                        office="Constitutional Amendment 2",
                        district_kind="na",
                        party="NP",
                        candidate=suffix,
                    )
                )
            continue

        for prefix, (office, district_kind) in PREFIX_CONTEST_MAP.items():
            if not h.startswith(prefix):
                continue
            suffix = h[len(prefix) :]
            if suffix in {"", "TOTAL", "TOT", "EST"}:
                break
            party = suffix
            candidate = "WRITE-IN" if party == "WI" else ""
            specs.append(
                ContestSpec(
                    column=header,
                    office=office,
                    district_kind=district_kind,
                    party=party,
                    candidate=candidate,
                )
            )
            break
    return specs


def district_from_row(kind: str, row: dict[str, str]) -> str:
    if kind == "na":
        return "NA"
    if kind == "cong":
        return normalize_numeric_district(row.get("CONGDIST", ""))
    if kind == "mnsen":
        return normalize_numeric_district(row.get("MNSENDIST", ""))
    if kind == "mnleg":
        return normalize_house_district(row.get("MNLEGDIST", ""))
    if kind == "unexpired":
        return "Unexpired Term"
    return "NA"


def party_aliases(party: str) -> list[str]:
    p = clean(party).upper()
    if p == "IP":
        return ["IP", "IND", "I"]
    if p == "IND":
        return ["IND", "IP", "I"]
    if p == "I":
        return ["I", "IND", "IP"]
    if p == "G":
        return ["G", "GP"]
    if p == "GP":
        return ["GP", "G"]
    return [p]


def lookup_candidate(
    candidate_map: dict[tuple[str, str, str], str],
    office: str,
    district: str,
    party: str,
) -> str:
    office = normalize_office(office)
    district = normalize_district_for_office(office, district)
    for p in party_aliases(party):
        hit = candidate_map.get((office, district, p), "")
        if hit:
            return hit
    if district != "NA":
        for p in party_aliases(party):
            hit = candidate_map.get((office, "NA", p), "")
            if hit:
                return hit
    return ""


def load_candidate_overrides(path: Path | None) -> dict[tuple[int, str, str, str], str]:
    overrides: dict[tuple[int, str, str, str], str] = {}
    if path is None or not path.exists():
        return overrides
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            year_text = clean(row.get("year", ""))
            if year_text == "":
                continue
            key = (
                int(year_text),
                normalize_office(row.get("office", "")),
                normalize_district_for_office(row.get("office", ""), row.get("district", "")),
                clean(row.get("party", "")).upper(),
            )
            candidate = clean(row.get("candidate", ""))
            if key[1] and key[3] and candidate:
                overrides[key] = candidate
    return overrides


def resolve_override(
    overrides: dict[tuple[int, str, str, str], str],
    year: int,
    office: str,
    district: str,
    party: str,
) -> str:
    office = normalize_office(office)
    district = normalize_district_for_office(office, district)
    p = clean(party).upper()
    for key in [
        (year, office, district, p),
        (year, office, "NA", p),
        (0, office, district, p),
        (0, office, "NA", p),
    ]:
        value = overrides.get(key, "")
        if value:
            return value
    return ""


def write_missing_report(counter: Counter[tuple[int, str, str, str]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["year", "office", "district", "party", "missing_rows"],
        )
        writer.writeheader()
        for (year, office, district, party), missing_rows in sorted(counter.items()):
            writer.writerow(
                {
                    "year": year,
                    "office": office,
                    "district": district,
                    "party": party,
                    "missing_rows": missing_rows,
                }
            )


def convert_job(
    job: SourceJob,
    candidate_overrides: dict[tuple[int, str, str, str], str],
    unknown_candidate_policy: str,
    missing_counter: Counter[tuple[int, str, str, str]],
) -> tuple[int, int]:
    if job.source_type == "xlsx":
        headers, rows = load_rows_from_xlsx(job.source_path)
    elif job.source_type == "csv":
        headers, rows = load_rows_from_csv(job.source_path)
    else:
        raise ValueError(f"Unsupported source type: {job.source_type}")

    contests = build_contests(headers)
    candidate_map = load_county_candidate_map(job.county_results_path)

    job.output_path.parent.mkdir(parents=True, exist_ok=True)
    out_rows = 0

    with job.output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["county", "precinct", "office", "district", "candidate", "party", "votes"],
        )
        writer.writeheader()

        for row in rows:
            county = clean(row.get("COUNTYNAME", ""))
            precinct = clean(row.get("PCTNAME", ""))
            if county == "" or precinct == "":
                continue

            for contest in contests:
                votes = parse_int(row.get(contest.column, ""))
                if votes is None:
                    continue

                district = district_from_row(contest.district_kind, row)
                office = contest.office
                party = contest.party.upper()
                candidate = contest.candidate

                if candidate == "":
                    candidate = lookup_candidate(candidate_map, office, district, party)

                if candidate == "":
                    candidate = resolve_override(candidate_overrides, job.year, office, district, party)

                if candidate == "":
                    if unknown_candidate_policy == "party_label":
                        candidate = f"{party} Candidate" if party else "Unknown Candidate"
                    elif unknown_candidate_policy == "unknown":
                        candidate = "Unknown Candidate"
                    else:
                        missing_counter[(job.year, office, district, party)] += 1

                writer.writerow(
                    {
                        "county": county,
                        "precinct": precinct,
                        "office": office,
                        "district": district,
                        "candidate": candidate,
                        "party": party,
                        "votes": votes,
                    }
                )
                out_rows += 1

    return len(rows), out_rows


def run_all(
    candidate_overrides: dict[tuple[int, str, str, str], str],
    unknown_candidate_policy: str,
    missing_counter: Counter[tuple[int, str, str, str]],
) -> None:
    for job in DEFAULT_JOBS:
        in_rows, out_rows = convert_job(
            job=job,
            candidate_overrides=candidate_overrides,
            unknown_candidate_policy=unknown_candidate_policy,
            missing_counter=missing_counter,
        )
        print(
            f"{job.year}: {job.source_path} -> {job.output_path} "
            f"({in_rows} source rows, {out_rows} output rows)"
        )


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert MN 2012–2024 precinct sources into OpenElections-style precinct CSVs."
    )
    parser.add_argument("--all", action="store_true", help="Run built-in jobs for 2012, 2014, 2016, 2018, and 2024.")
    parser.add_argument(
        "--candidate-overrides",
        type=Path,
        default=None,
        help="Optional override CSV with year,office,district,party,candidate.",
    )
    parser.add_argument(
        "--unknown-candidate-policy",
        choices=["blank", "party_label", "unknown"],
        default="blank",
        help="How to handle unresolved candidate names.",
    )
    parser.add_argument(
        "--missing-report",
        type=Path,
        default=Path("Data/missing_candidates_modern_precincts.csv"),
        help="Output report of remaining missing candidate names.",
    )
    args = parser.parse_args()

    if not args.all:
        parser.error("Use --all for this script.")

    overrides = load_candidate_overrides(args.candidate_overrides)
    missing_counter: Counter[tuple[int, str, str, str]] = Counter()

    run_all(
        candidate_overrides=overrides,
        unknown_candidate_policy=args.unknown_candidate_policy,
        missing_counter=missing_counter,
    )

    if args.unknown_candidate_policy == "blank":
        write_missing_report(missing_counter, args.missing_report)
        print(f"Missing-candidate report written: {args.missing_report} ({sum(missing_counter.values())} rows)")


if __name__ == "__main__":
    main()
